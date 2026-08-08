# -*- coding: utf-8 -*-
"""Build 发货费用明细 (第16批 format) from 第15次 packing boxes + unique-code exports."""
from __future__ import annotations

import shutil
from collections import defaultdict
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

FEE_TEMPLATE = Path(r"f:\第16批手办发货费用明细.xlsx")
BAK_INV = Path(
    r"c:\Users\Aick\AppData\Roaming\kingsoft\office6\templates\et\zh_CN"
    r"\第15次手办发香港.xlsx.bak-inv-packing"
)
WPS_DST = Path(
    r"c:\Users\Aick\AppData\Roaming\kingsoft\office6\templates\et\zh_CN\第15次手办发香港.xlsx"
)
OUT_PATHS = [
    Path(r"f:\第15次手办发货费用明细_完善.xlsx"),
    WPS_DST.with_name("第15次手办发货费用明细_完善.xlsx"),
]
UNIQUE_FILES = [
    Path(r"f:\唯一码库存明细_20260804191230.xlsx"),
    Path(r"f:\唯一码库存明细_20260804190921.xlsx"),
    Path(r"f:\唯一码库存明细_20260804190908.xlsx"),
    Path(r"f:\唯一码库存明细_20260804190835.xlsx"),
]

DETAIL_COLS = [
    "商品图片",
    "商品唯一码",
    "SKU编码",
    "货号",
    "商品名",
    "商品分类",
    "规格",
    "条形码",
    "批次号",
    "供应商",
    "采购价",
    "质量类型",
    "质量等级",
    "在仓状态",
    "数量",
    "当前所在仓库",
    "当前所在库区",
    "当前所在位置",
    "首次入库时间",
    "库存类型",
    "库龄",
]


class TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._table: list[list[str]] | None = None
        self._row: list[str] | None = None
        self._cell: str | None = None
        self._in_cell = False

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag == "table":
            self._table = []
        elif tag == "tr" and self._table is not None:
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = ""
            self._in_cell = True

    def handle_endtag(self, tag: str) -> None:
        if tag in ("td", "th") and self._in_cell:
            assert self._row is not None and self._cell is not None
            self._row.append(self._cell.strip())
            self._cell = None
            self._in_cell = False
        elif tag == "tr" and self._row is not None:
            assert self._table is not None
            self._table.append(self._row)
            self._row = None
        elif tag == "table" and self._table is not None:
            self.tables.append(self._table)
            self._table = None

    def handle_data(self, data: str) -> None:
        if self._in_cell and self._cell is not None:
            self._cell += data


def parse_html_table(path: Path) -> list[list[str]]:
    parser = TableParser()
    parser.feed(path.read_text(encoding="utf-8"))
    return parser.tables[0]


def normalize_barcode(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def load_unique_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for path in UNIQUE_FILES:
        table = parse_html_table(path)
        header_idx = next(
            i for i, row in enumerate(table) if "条形码" in row or "商品唯一码" in row
        )
        headers = table[header_idx]
        for raw in table[header_idx + 1 :]:
            item = {
                headers[i]: (raw[i] if i < len(raw) else "")
                for i in range(len(headers))
            }
            item["_source"] = path.name
            rows.append(item)
        print(f"{path.name}: {len(table) - header_idx - 1} rows")
    return rows


def load_packing_and_prices() -> tuple[dict[str, dict], dict[str, float], str]:
    tmp = Path(__file__).with_name("_tmp_bak.xlsx")
    shutil.copy2(BAK_INV, tmp)
    wb = load_workbook(tmp)
    packing: dict[str, dict] = {}
    ws_p = wb["PACKING LIST "]
    for r in range(22, ws_p.max_row + 1):
        pack_no = ws_p.cell(r, 1).value
        if not pack_no or str(pack_no).upper() == "TOTAL":
            continue
        packing[str(pack_no).strip()] = {
            "qty": ws_p.cell(r, 3).value,
            "net": ws_p.cell(r, 4).value,
            "gross": ws_p.cell(r, 5).value,
            "volume": ws_p.cell(r, 6).value,
        }
    prices: dict[str, float] = {}
    ws_i = wb["inv"]
    invoice_no = str(ws_i["F3"].value or "")
    for r in range(8, ws_i.max_row + 1):
        code = normalize_barcode(ws_i.cell(r, 1).value)
        price = ws_i.cell(r, 6).value
        if code and price is not None:
            prices[code] = float(price)
    tmp.unlink(missing_ok=True)
    return packing, prices, invoice_no


def main() -> None:
    unique_rows = load_unique_rows()
    packing, prices, invoice_no = load_packing_and_prices()
    print("packing boxes:", packing)
    print("prices:", prices)
    print("invoice:", invoice_no)

    # location order from packing list
    loc_order = list(packing.keys())
    for loc in sorted({u.get("当前所在位置", "") for u in unique_rows}):
        if loc and loc not in loc_order:
            loc_order.append(loc)

    # group unique by (location, barcode)
    groups: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for u in unique_rows:
        loc = u.get("当前所在位置", "") or ""
        bc = normalize_barcode(u.get("条形码"))
        groups[(loc, bc)].append(u)

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True)

    # Summary labels aligned with 第16批手办发货费用明细
    tmpl = load_workbook(FEE_TEMPLATE)
    tmpl_ws = tmpl.active
    summary_labels: list[str] = []
    for r in range(1, tmpl_ws.max_row + 1):
        e = tmpl_ws.cell(r, 5).value
        b = tmpl_ws.cell(r, 2).value
        # label rows sit below data: col E text, no 箱号
        if isinstance(e, str) and b is None and e != "数量":
            summary_labels.append(e)
    if len(summary_labels) < 3:
        summary_labels = [
            "商品价值",
            "运费汇率",
            "报关费用",
            "运费单价（JPY）",
            "已支付运费（CNY）",
            "已支付报关（CNY）",
            "已支付总额（CNY）",
        ]

    out = Workbook()
    ws = out.active
    ws.title = "发货费用明细"
    headers = ["单号", "箱号", "品名", "条形码", "数量", "合计JPY", "净重", "毛重", "体积", "装箱号", "唯一码"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.border = thin

    # Build fee rows: one per location+barcode, box_no sequential by location
    fee_lines: list[dict] = []
    box_no_of_loc = {loc: i for i, loc in enumerate(loc_order, start=1)}
    for loc in loc_order:
        loc_groups = [(k, v) for k, v in groups.items() if k[0] == loc]
        # stable order by barcode
        loc_groups.sort(key=lambda kv: kv[0][1])
        first = True
        pack = packing.get(loc, {})
        for (_loc, bc), items in loc_groups:
            name = items[0].get("商品名") or ""
            qty = len(items)
            unit = prices.get(bc)
            amount = round(unit * qty, 2) if unit is not None else None
            codes = [i.get("商品唯一码", "") for i in items]
            fee_lines.append(
                {
                    "box_no": box_no_of_loc[loc],
                    "pack_no": loc,
                    "name": name,
                    "barcode": bc,
                    "qty": qty,
                    "amount_jpy": amount,
                    "net": pack.get("net") if first else None,
                    "gross": pack.get("gross") if first else None,
                    "volume": pack.get("volume") if first else None,
                    "codes": codes,
                    "items": items,
                }
            )
            first = False

    # Write fee rows
    for i, line in enumerate(fee_lines, start=1):
        r = i + 1
        values = [
            invoice_no if i == 1 else None,  # merged later
            line["box_no"],
            line["name"],
            line["barcode"],
            line["qty"],
            line["amount_jpy"],
            line["net"],
            line["gross"],
            line["volume"],
            line["pack_no"],
            " / ".join(line["codes"]),
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(r, c, v)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    last_data_row = 1 + len(fee_lines)
    if len(fee_lines) > 1:
        ws.merge_cells(start_row=2, start_column=1, end_row=last_data_row, end_column=1)
        ws.cell(2, 1).alignment = Alignment(vertical="center", wrap_text=True)

    # totals
    total_row = last_data_row + 1
    total_qty = sum(int(x["qty"]) for x in fee_lines)
    total_jpy = sum(float(x["amount_jpy"] or 0) for x in fee_lines)
    ws.cell(total_row, 5, total_qty).border = thin
    ws.cell(total_row, 6, total_jpy).border = thin
    ws.cell(total_row, 5).font = header_font
    ws.cell(total_row, 6).font = header_font

    # summary labels (empty values for fill-in)
    for i, label in enumerate(summary_labels):
        r = total_row + 1 + i
        ws.cell(r, 5, label)

    for c, w in enumerate([18, 8, 28, 16, 8, 12, 8, 8, 12, 12, 40], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Sheet: 箱内唯一码明细
    ws_d = out.create_sheet("箱内唯一码明细")
    d_headers = ["箱号", "装箱号", "费用表品名", "费用表条形码", "合计JPY"] + DETAIL_COLS + ["来源文件"]
    for c, h in enumerate(d_headers, 1):
        cell = ws_d.cell(1, c, h)
        cell.font = header_font
        cell.border = thin
    dr = 2
    for line in fee_lines:
        for u in line["items"]:
            mapping = {
                "箱号": line["box_no"],
                "装箱号": line["pack_no"],
                "费用表品名": line["name"],
                "费用表条形码": line["barcode"],
                "合计JPY": line["amount_jpy"],
                "来源文件": u.get("_source", ""),
            }
            for h in DETAIL_COLS:
                mapping[h] = u.get(h, "")
            for c, h in enumerate(d_headers, 1):
                cell = ws_d.cell(dr, c, mapping.get(h, ""))
                cell.border = thin
            dr += 1

    # Sheet: 按装箱号
    ws_b = out.create_sheet("按装箱号")
    b_headers = ["箱号", "装箱号", "装箱数量(仓)", "净重", "毛重", "体积", "唯一码件数", "SKU数", "品名汇总"]
    for c, h in enumerate(b_headers, 1):
        cell = ws_b.cell(1, c, h)
        cell.font = header_font
        cell.border = thin
    br = 2
    for loc in loc_order:
        lines = [x for x in fee_lines if x["pack_no"] == loc]
        pack = packing.get(loc, {})
        names = "；".join(f"{x['name']}×{x['qty']}" for x in lines)
        vals = [
            box_no_of_loc[loc],
            loc,
            pack.get("qty"),
            pack.get("net"),
            pack.get("gross"),
            pack.get("volume"),
            sum(x["qty"] for x in lines),
            len(lines),
            names,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws_b.cell(br, c, v)
            cell.border = thin
        br += 1

    for sheet in (ws_d, ws_b):
        for c in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(c)].width = 14

    # Save outputs
    primary = OUT_PATHS[0]
    out.save(primary)
    print(f"fee lines={len(fee_lines)} unique={len(unique_rows)} total_qty={total_qty} total_jpy={total_jpy}")
    print("saved", primary)

    for path in OUT_PATHS[1:]:
        shutil.copy2(primary, path)
        print("copied", path)

    # Overwrite WPS template with enriched fee-detail workbook
    shutil.copy2(primary, WPS_DST)
    print("WPS overwritten", WPS_DST)

    # Also keep a clean 第16批 shell copy in templates folder
    shell = WPS_DST.with_name("第16批手办发货费用明细_模板.xlsx")
    shutil.copy2(FEE_TEMPLATE, shell)
    print("template shell", shell)


if __name__ == "__main__":
    main()
