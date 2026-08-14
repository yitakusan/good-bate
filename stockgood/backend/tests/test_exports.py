from __future__ import annotations

import hashlib
import unittest
from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.database import DATA_DIR
from app.services import outbound_batches as outbound_svc
from app.services.inv_template import (
    CONSIGNEE_EN,
    INV_SHEET,
    PACK_SHEET,
    TEMPLATE_PATH,
    format_inv_no,
)
from factories import make_packed_outbound_batch, make_priced_in_stock_order
from harness import IsolatedDbTestCase

FEE_TEMPLATE = DATA_DIR / "templates" / "fee_detail.xlsx"
FEE_SHEET = "发货费用明细"
FEE_ORDERS_SHEET = "对应订单"


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


class ExportTemplateContractTests(IsolatedDbTestCase):
    def test_inv_fills_packing_f5_not_f6_and_does_not_rewrite_template(self) -> None:
        before = _sha256(TEMPLATE_PATH)
        template = load_workbook(TEMPLATE_PATH)
        self.assertIn(INV_SHEET, template.sheetnames)
        self.assertIn(PACK_SHEET, template.sheetnames)
        sample_f5 = template[PACK_SHEET]["F5"].value
        sample_f6 = template[PACK_SHEET]["F6"].value

        order = make_priced_in_stock_order(order_ref="INV-001")
        batch = make_packed_outbound_batch(order, tracking_no="INV-TRK-001")
        content, filename = outbound_svc.export_inv_xlsx(batch["id"])
        self.assertEqual(content[:2], b"PK")
        total_qty = sum(int(line["qty"]) for line in order["lines"])
        expected_no = format_inv_no(date(2026, 8, 14), batch["id"], total_qty)
        self.assertTrue(filename.startswith("INV_20260814_FIT"))
        self.assertTrue(filename.endswith(".xlsx"))

        wb = load_workbook(BytesIO(content))
        self.assertEqual(wb.sheetnames[0], INV_SHEET)
        self.assertEqual(wb.sheetnames[1], PACK_SHEET)
        self.assertEqual(PACK_SHEET, "PACKING LIST ")
        self.assertTrue(PACK_SHEET.endswith(" "))

        inv = wb[INV_SHEET]
        pack = wb[PACK_SHEET]
        self.assertEqual(inv["F3"].value, expected_no)
        self.assertEqual(inv["F4"].value, expected_no)
        self.assertEqual(pack["F5"].value, expected_no)
        self.assertNotEqual(pack["F5"].value, sample_f5)
        self.assertEqual(pack["F6"].value, sample_f6)
        self.assertNotEqual(pack["F6"].value, expected_no)
        self.assertEqual(inv["A3"].value, CONSIGNEE_EN)
        self.assertEqual(pack["A8"].value, CONSIGNEE_EN)
        self.assertEqual(inv.cell(8, 1).value, "111")
        self.assertEqual(pack.cell(22, 6).value, "10*8*6")

        self.assertEqual(_sha256(TEMPLATE_PATH), before)

    def test_inv_preview_uses_batch_zero_in_invoice_number(self) -> None:
        order = make_priced_in_stock_order(order_ref="INV-PREVIEW")
        item_ids = [line["id"] for line in order["lines"]]
        content, filename = outbound_svc.export_inv_preview_xlsx(
            {
                "boxes": [
                    {
                        "box_no": 1,
                        "item_ids": item_ids,
                        "net_weight": 1.0,
                        "gross_weight": 1.2,
                        "length_cm": 10.0,
                        "width_cm": 8.0,
                        "height_cm": 6.0,
                    }
                ],
                "invoice_ship_date": "2026-08-14",
            }
        )
        total_qty = sum(int(line["qty"]) for line in order["lines"])
        expected_no = format_inv_no(date(2026, 8, 14), 0, total_qty)
        wb = load_workbook(BytesIO(content))
        self.assertEqual(wb[PACK_SHEET]["F5"].value, expected_no)
        self.assertIsNone(wb[PACK_SHEET]["F6"].value)
        self.assertIn("FIT0_", filename)

    def test_fee_detail_keeps_dual_sheets_headers_and_order_rows(self) -> None:
        before = _sha256(FEE_TEMPLATE)
        template = load_workbook(FEE_TEMPLATE)
        self.assertIn(FEE_SHEET, template.sheetnames)
        self.assertIn(FEE_ORDERS_SHEET, template.sheetnames)
        headers = [template[FEE_SHEET].cell(1, c).value for c in range(1, 13)]

        order = make_priced_in_stock_order(order_ref="FEE-001")
        batch = make_packed_outbound_batch(order, tracking_no="FEE-TRK-001")
        raw = outbound_svc.export_fee_detail_xlsx(batch["id"])
        self.assertEqual(raw[:2], b"PK")

        wb = load_workbook(BytesIO(raw))
        self.assertIn(FEE_SHEET, wb.sheetnames)
        self.assertIn(FEE_ORDERS_SHEET, wb.sheetnames)
        ws = wb[FEE_SHEET]
        exported_headers = [ws.cell(1, c).value for c in range(1, 13)]
        self.assertEqual(exported_headers, headers)
        self.assertEqual(headers[1], "箱号")
        self.assertEqual(headers[2], "订单号")
        self.assertEqual(headers[3], "品名")
        self.assertEqual(ws.cell(2, 1).value, f"OB-{batch['id']}")
        self.assertEqual(ws.cell(2, 2).value, 1)
        self.assertEqual(ws.cell(2, 3).value, "FEE-001")

        orders_ws = wb[FEE_ORDERS_SHEET]
        self.assertEqual(orders_ws.cell(2, 1).value, 1)
        self.assertEqual(orders_ws.cell(2, 3).value, "FEE-001")
        self.assertEqual(int(orders_ws.cell(2, 4).value), 3)

        self.assertEqual(_sha256(FEE_TEMPLATE), before)


if __name__ == "__main__":
    unittest.main()
