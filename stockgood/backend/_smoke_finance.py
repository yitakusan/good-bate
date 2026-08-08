# -*- coding: utf-8 -*-
"""Shadow-DB smoke for finance: rate, outbound receivable, Excel, month summary."""
from __future__ import annotations

import os
from io import BytesIO

os.environ["STOCKGOOD_DB_MODE"] = "shadow"

from openpyxl import load_workbook

from app.database import get_conn, get_db_path, init_db
from app.models import LineCreate, OrderCreate
from app.services import finance as finance_svc
from app.services import orders as orders_svc
from app.services import outbound_batches as outbound_svc


def main() -> None:
    print("DB:", get_db_path())
    assert "shadow" in str(get_db_path()), "must use shadow DB"
    init_db()

    with get_conn() as conn:
        refs = conn.execute(
            "SELECT id FROM orders WHERE order_ref LIKE 'FIN-SMOKE%'"
        ).fetchall()
        for row in refs:
            conn.execute("DELETE FROM orders WHERE id = ?", (row["id"],))
        ships = conn.execute(
            "SELECT id, batch_id FROM shipments WHERE tracking_no LIKE 'FIN-TRK%'"
        ).fetchall()
        batch_ids = {s["batch_id"] for s in ships if s["batch_id"]}
        for ship in ships:
            conn.execute("DELETE FROM shipments WHERE id = ?", (ship["id"],))
        for bid in batch_ids:
            conn.execute("DELETE FROM outbound_batches WHERE id = ?", (bid,))
        conn.execute(
            "DELETE FROM outbound_batches WHERE note LIKE 'FIN-SMOKE%'"
        )

    order = orders_svc.create_order(
        OrderCreate(
            order_ref="FIN-SMOKE-001",
            shop="test-shop",
            shipping_fee=500,
            exchange_rate=0.05,
            lines=[
                LineCreate(
                    name="Smoke Item A", qty=2, unit_cost=1000, barcode="111"
                ),
                LineCreate(
                    name="Smoke Item B", qty=1, unit_cost=2000, barcode="222"
                ),
            ],
        )
    )
    print(
        "order",
        order["id"],
        "goods_cny",
        order["goods_total_cny"],
        "total_cny",
        order["order_total_cny"],
    )
    assert abs(order["goods_total_cny"] - 200.0) < 1e-9
    assert abs(order["order_total_cny"] - 225.0) < 1e-9

    item_ids = [line["id"] for line in order["lines"]]
    with get_conn() as conn:
        for iid in item_ids:
            conn.execute(
                "UPDATE items SET status = 'in_stock' WHERE id = ?", (iid,)
            )
        conn.execute(
            "UPDATE orders SET status = 'in_stock' WHERE id = ?", (order["id"],)
        )

    batch = outbound_svc.create_batch(
        boxes=[
            {
                "box_no": 1,
                "carrier": "yamato",
                "tracking_no": "FIN-TRK-OUT-001",
                "item_ids": item_ids,
            }
        ],
        note="FIN-SMOKE batch",
        freight_exchange_rate=0.048,
        freight_unit_price_jpy=1000,
        chargeable_weight=2.5,
    )
    print(
        "batch",
        batch["id"],
        "goods_cny",
        batch["goods_receivable_cny"],
        "freight",
        batch["freight_cny"],
        "recv",
        batch["amount_receivable_cny"],
        batch["payment_status"],
    )
    assert abs(batch["goods_receivable_cny"] - 225.0) < 1e-9
    assert abs(batch["freight_cny"] - 120.0) < 1e-9
    assert abs(batch["amount_receivable_cny"] - 345.0) < 1e-9
    assert batch["payment_status"] == "unpaid"

    partial = outbound_svc.update_finance(batch["id"], {"amount_received_cny": 200})
    assert partial["payment_status"] == "partial"
    assert abs(partial["amount_unreceived_cny"] - 145.0) < 1e-9

    paid = outbound_svc.update_finance(batch["id"], {"amount_received_cny": 345})
    assert paid["payment_status"] == "paid"

    xlsx = outbound_svc.export_fee_detail_xlsx(batch["id"])
    assert xlsx[:2] == b"PK"
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, 13)]
    print("headers", headers)
    assert headers[1] == "箱号"
    assert headers[2] == "订单号"
    assert headers[3] == "品名"
    assert "下单汇率" in headers
    assert ws.cell(2, 3).value == "FIN-SMOKE-001"

    summary = finance_svc.month_summary()
    print(
        "finance",
        summary["month"],
        "ordered_cny",
        summary["ordered"]["total_cny"],
        "outbound_recv",
        summary["outbound"]["amount_receivable_cny"],
    )
    print("SMOKE OK")


if __name__ == "__main__":
    main()
