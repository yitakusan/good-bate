from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Optional

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from app.database import DATA_DIR, get_conn
from app.services import action_log
from app.services.order_status import sync_order_status
from app.services.shipments import _shipment_with_items
from app.services.stock_boxes import release_orders as release_stock_box_orders


def _now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _as_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _require_shared_batch_tracking(boxes: list[dict[str, Any]]) -> str:
    filled = [(box.get("tracking_no") or "").strip() for box in boxes]
    nonempty = [t for t in filled if t]
    if not nonempty:
        raise HTTPException(status_code=400, detail="出库批次需要运单号")
    if len(set(nonempty)) > 1:
        raise HTTPException(
            status_code=400,
            detail="同一出库批次必须使用同一个运单号",
        )
    return nonempty[0]


def _payment_status(receivable: Optional[float], received: float) -> str:
    if receivable is None:
        return "unpaid" if received <= 0 else "partial"
    if received <= 0:
        return "unpaid"
    if received + 1e-9 >= receivable:
        return "paid"
    return "partial"


def _compute_freight_cny(
    unit_price: Optional[float],
    weight: Optional[float],
    rate: Optional[float],
) -> Optional[float]:
    if unit_price is None or weight is None or rate is None:
        return None
    if rate <= 0:
        return None
    return round(float(unit_price) * float(weight) * float(rate), 2)


def _recompute_batch_totals(
    goods_receivable_cny: Optional[float],
    freight_cny: Optional[float],
    amount_received_cny: float,
) -> tuple[Optional[float], str, Optional[float]]:
    parts = [v for v in (goods_receivable_cny, freight_cny) if v is not None]
    receivable = round(sum(parts), 2) if parts else None
    status = _payment_status(receivable, amount_received_cny)
    unreceived = (
        round(receivable - amount_received_cny, 2) if receivable is not None else None
    )
    return receivable, status, unreceived


def _box_out(conn, shipment_id: int) -> dict[str, Any]:
    ship = _shipment_with_items(conn, shipment_id)
    return {
        "id": ship["id"],
        "batch_id": ship["batch_id"],
        "box_no": ship["box_no"],
        "carrier": ship["carrier"],
        "tracking_no": ship["tracking_no"],
        "tracking_url": ship["tracking_url"],
        "status": ship["status"],
        "shipped_at": ship["shipped_at"],
        "delivered_at": ship["delivered_at"],
        "note": ship.get("note") or "",
        "net_weight": _as_float(ship.get("net_weight")),
        "gross_weight": _as_float(ship.get("gross_weight")),
        "length_cm": _as_float(ship.get("length_cm")),
        "width_cm": _as_float(ship.get("width_cm")),
        "height_cm": _as_float(ship.get("height_cm")),
        "items": ship["items"],
        "order_groups": ship["order_groups"],
    }


def _batch_out(conn, batch_id: int) -> dict[str, Any]:
    batch = conn.execute(
        "SELECT * FROM outbound_batches WHERE id = ?", (batch_id,)
    ).fetchone()
    if not batch:
        raise HTTPException(status_code=404, detail="outbound batch not found")
    boxes = conn.execute(
        """
        SELECT id FROM shipments
        WHERE batch_id = ? AND direction = 'outbound'
        ORDER BY box_no, id
        """,
        (batch_id,),
    ).fetchall()
    box_outs = [_box_out(conn, b["id"]) for b in boxes]
    received = _as_float(batch["amount_received_cny"]) or 0.0
    goods_cny = _as_float(batch["goods_receivable_cny"])
    freight_cny = _as_float(batch["freight_cny"])
    receivable = _as_float(batch["amount_receivable_cny"])
    if receivable is None:
        receivable, _, unreceived = _recompute_batch_totals(
            goods_cny, freight_cny, received
        )
    else:
        unreceived = round(receivable - received, 2)
    status = batch["payment_status"] or _payment_status(receivable, received)
    return {
        "id": batch["id"],
        "note": batch["note"],
        "created_at": batch["created_at"],
        "boxes": box_outs,
        "box_count": len(box_outs),
        "item_count": sum(len(b["items"]) for b in box_outs),
        "goods_jpy": _as_float(batch["goods_jpy"]),
        "order_shipping_jpy": _as_float(batch["order_shipping_jpy"]),
        "goods_receivable_cny": goods_cny,
        "freight_exchange_rate": _as_float(batch["freight_exchange_rate"]),
        "freight_unit_price_jpy": _as_float(batch["freight_unit_price_jpy"]),
        "chargeable_weight": _as_float(batch["chargeable_weight"]),
        "freight_cny": freight_cny,
        "amount_receivable_cny": receivable,
        "amount_received_cny": received,
        "amount_unreceived_cny": unreceived,
        "payment_status": status,
        "payment_note": batch["payment_note"] or "",
        "invoice_ship_date": (batch["invoice_ship_date"] if "invoice_ship_date" in batch.keys() else None)
        or None,
    }


def list_batches(limit: int = 50) -> list[dict[str, Any]]:
    limit = max(1, min(limit, 200))
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id FROM outbound_batches ORDER BY id DESC LIMIT ?",
            (limit,),
        ).fetchall()
        return [_batch_out(conn, r["id"]) for r in rows]


def get_batch(batch_id: int) -> dict[str, Any]:
    with get_conn() as conn:
        return _batch_out(conn, batch_id)


def _lock_receivables(
    conn,
    item_ids: set[int],
    order_ids: set[int],
) -> dict[str, Optional[float]]:
    """Lock goods JPY/CNY at outbound create time using each order's rate."""
    goods_jpy = 0.0
    goods_cny = 0.0
    order_shipping_jpy = 0.0
    order_shipping_cny = 0.0
    has_goods_priced = False
    has_goods_cny = False
    has_ship_cny = False

    placeholders = ",".join("?" * len(item_ids))
    lines = conn.execute(
        f"""
        SELECT i.id, i.order_id, i.unit_cost, i.qty, o.exchange_rate
        FROM items i
        JOIN orders o ON o.id = i.order_id
        WHERE i.id IN ({placeholders})
        """,
        list(item_ids),
    ).fetchall()
    for line in lines:
        if line["unit_cost"] is None:
            continue
        line_jpy = float(line["unit_cost"]) * int(line["qty"])
        goods_jpy += line_jpy
        has_goods_priced = True
        rate = _as_float(line["exchange_rate"])
        if rate is not None and rate > 0:
            goods_cny += line_jpy * rate
            has_goods_cny = True

    for oid in order_ids:
        order = conn.execute(
            "SELECT shipping_fee, exchange_rate FROM orders WHERE id = ?",
            (oid,),
        ).fetchone()
        if not order:
            continue
        ship = _as_float(order["shipping_fee"]) or 0.0
        order_shipping_jpy += ship
        rate = _as_float(order["exchange_rate"])
        if rate is not None and rate > 0:
            order_shipping_cny += ship * rate
            has_ship_cny = True

    goods_receivable = None
    if has_goods_cny or has_ship_cny:
        goods_receivable = round(
            (goods_cny if has_goods_cny else 0.0)
            + (order_shipping_cny if has_ship_cny else 0.0),
            2,
        )

    return {
        "goods_jpy": round(goods_jpy, 2) if has_goods_priced else None,
        "order_shipping_jpy": round(order_shipping_jpy, 2),
        "goods_receivable_cny": goods_receivable,
    }


def create_batch(
    boxes: list[dict[str, Any]],
    note: str = "",
    *,
    allow_missing_barcode: bool = False,
    missing_barcode_note: str = "",
    freight_exchange_rate: Optional[float] = None,
    freight_unit_price_jpy: Optional[float] = None,
    chargeable_weight: Optional[float] = None,
    invoice_ship_date: Optional[str] = None,
) -> dict[str, Any]:
    """
    Outbound batch: multiple boxes, each with independent box_no + tracking.
    A box may contain lines from multiple orders (secondary group by order).
    No partial outbound: if any line of an order is included, all in_stock
    lines of that order must be in this batch (any boxes).
    Barcode required on every line unless allow_missing_barcode + note.
    Locks goods receivable CNY at create time; international freight uses
    an independent rate.
    """
    if not boxes:
        raise HTTPException(status_code=400, detail="boxes required")
    special_note = (missing_barcode_note or "").strip()
    if allow_missing_barcode and not special_note:
        raise HTTPException(
            status_code=400,
            detail="勾选特殊情况时必须填写无条形码备注",
        )

    for label, value in (
        ("freight_exchange_rate", freight_exchange_rate),
        ("freight_unit_price_jpy", freight_unit_price_jpy),
        ("chargeable_weight", chargeable_weight),
    ):
        if value is not None and float(value) < 0:
            raise HTTPException(status_code=400, detail=f"{label} invalid")
    if freight_exchange_rate is not None and float(freight_exchange_rate) <= 0:
        raise HTTPException(status_code=400, detail="freight_exchange_rate must be > 0")

    normalized: list[dict[str, Any]] = []
    seen_items: set[int] = set()
    used_box_nos: set[int] = set()
    batch_tracking = _require_shared_batch_tracking(boxes)

    for i, box in enumerate(boxes, start=1):
        ids = list(box.get("item_ids") or [])
        if not ids:
            raise HTTPException(status_code=400, detail=f"第{i}箱没有明细行")
        unique: list[int] = []
        for iid in ids:
            if iid in seen_items:
                raise HTTPException(
                    status_code=400, detail=f"明细行 #{iid} 被分到多个箱"
                )
            seen_items.add(iid)
            unique.append(iid)

        carrier = (box.get("carrier") or "other").strip().lower()
        if carrier not in ("yamato", "sagawa", "other"):
            raise HTTPException(status_code=400, detail=f"第{i}箱承运商无效")

        box_no = box.get("box_no")
        if box_no is None:
            box_no = i
        box_no = int(box_no)
        if box_no < 1:
            raise HTTPException(status_code=400, detail=f"第{i}箱箱号无效")
        if box_no in used_box_nos:
            raise HTTPException(status_code=400, detail=f"箱号重复: {box_no}")
        used_box_nos.add(box_no)

        normalized.append(
            {
                "box_no": box_no,
                "carrier": carrier,
                "tracking_no": batch_tracking,
                "note": (box.get("note") or "").strip(),
                "item_ids": unique,
                "net_weight": _as_float(box.get("net_weight")),
                "gross_weight": _as_float(box.get("gross_weight")),
                "length_cm": _as_float(box.get("length_cm")),
                "width_cm": _as_float(box.get("width_cm")),
                "height_cm": _as_float(box.get("height_cm")),
            }
        )

    with get_conn() as conn:
        exists = conn.execute(
            "SELECT id FROM shipments WHERE tracking_no = ?",
            (batch_tracking,),
        ).fetchone()
        if exists:
            raise HTTPException(
                status_code=409,
                detail=f"运单号已被其他出库批次或进库使用: {batch_tracking}",
            )

        placeholders = ",".join("?" * len(seen_items))
        rows = conn.execute(
            f"""
            SELECT id, order_id, order_ref, status, qty, name, barcode
            FROM items WHERE id IN ({placeholders})
            """,
            list(seen_items),
        ).fetchall()
        by_id = {r["id"]: r for r in rows}
        if len(by_id) != len(seen_items):
            missing = [iid for iid in seen_items if iid not in by_id]
            raise HTTPException(status_code=404, detail=f"items not found: {missing}")

        no_barcode = [
            r
            for r in by_id.values()
            if not (r["barcode"] or "").strip()
        ]
        if no_barcode and not allow_missing_barcode:
            names = "、".join(f"「{r['name']}」" for r in no_barcode[:5])
            more = f" 等 {len(no_barcode)} 行" if len(no_barcode) > 5 else ""
            raise HTTPException(
                status_code=400,
                detail=(
                    f"出库必须登记条形码，以下货品尚未填写：{names}{more}。"
                    "请补条码，或勾选特殊情况并备注。"
                ),
            )

        for iid, row in by_id.items():
            if row["status"] != "in_stock":
                raise HTTPException(
                    status_code=400,
                    detail=f"「{row['name']}」不是在库，不能出库",
                )
            if row["order_id"] is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"明细行 #{iid} 未关联订单",
                )

        already = conn.execute(
            f"""
            SELECT item_id FROM shipment_items
            JOIN shipments s ON s.id = shipment_items.shipment_id
            WHERE item_id IN ({placeholders}) AND s.direction = 'outbound'
            """,
            list(seen_items),
        ).fetchall()
        if already:
            ids = [r["item_id"] for r in already]
            raise HTTPException(
                status_code=400, detail=f"items already in outbound shipment: {ids}"
            )

        # No partial outbound: every in_stock line of touched orders must be included
        order_ids = {int(by_id[iid]["order_id"]) for iid in seen_items}
        for oid in order_ids:
            stock_rows = conn.execute(
                """
                SELECT id FROM items
                WHERE order_id = ? AND status = 'in_stock'
                """,
                (oid,),
            ).fetchall()
            stock_ids = {r["id"] for r in stock_rows}
            missing = stock_ids - seen_items
            if missing:
                ref = conn.execute(
                    "SELECT order_ref FROM orders WHERE id = ?", (oid,)
                ).fetchone()
                label = (ref["order_ref"] if ref else "") or f"#{oid}"
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"订单 {label} 不能部分出库，未装箱的在库明细: "
                        f"{sorted(missing)}"
                    ),
                )

        batch_note = note.strip()
        if allow_missing_barcode and no_barcode:
            tag = f"【无条码特批】{special_note}"
            batch_note = f"{batch_note} | {tag}".strip(" |") if batch_note else tag

        locked = _lock_receivables(conn, seen_items, order_ids)
        freight_rate = (
            float(freight_exchange_rate) if freight_exchange_rate is not None else None
        )
        freight_unit = (
            float(freight_unit_price_jpy)
            if freight_unit_price_jpy is not None
            else None
        )
        freight_weight = (
            float(chargeable_weight) if chargeable_weight is not None else None
        )
        freight_cny = _compute_freight_cny(freight_unit, freight_weight, freight_rate)
        receivable, pay_status, _ = _recompute_batch_totals(
            locked["goods_receivable_cny"], freight_cny, 0.0
        )

        cur = conn.execute(
            """
            INSERT INTO outbound_batches (
                note, created_at,
                goods_jpy, order_shipping_jpy, goods_receivable_cny,
                freight_exchange_rate, freight_unit_price_jpy, chargeable_weight,
                freight_cny, amount_receivable_cny, amount_received_cny,
                payment_status, payment_note, invoice_ship_date
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, '', ?)
            """,
            (
                batch_note,
                _now(),
                locked["goods_jpy"],
                locked["order_shipping_jpy"],
                locked["goods_receivable_cny"],
                freight_rate,
                freight_unit,
                freight_weight,
                freight_cny,
                receivable,
                pay_status,
                (invoice_ship_date or "").strip() or None,
            ),
        )
        batch_id = int(cur.lastrowid)
        shipment_ids: list[int] = []
        item_restores: list[dict[str, Any]] = []
        order_restores = {
            oid: conn.execute(
                "SELECT status FROM orders WHERE id = ?", (oid,)
            ).fetchone()["status"]
            for oid in order_ids
        }

        for box in sorted(normalized, key=lambda b: b["box_no"]):
            scur = conn.execute(
                """
                INSERT INTO shipments (
                    direction, carrier, tracking_no, shipped_at, status,
                    batch_id, box_no, note,
                    net_weight, gross_weight, length_cm, width_cm, height_cm
                ) VALUES ('outbound', ?, ?, ?, 'shipped', ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    box["carrier"],
                    box["tracking_no"],
                    _now(),
                    batch_id,
                    box["box_no"],
                    box.get("note") or "",
                    box.get("net_weight"),
                    box.get("gross_weight"),
                    box.get("length_cm"),
                    box.get("width_cm"),
                    box.get("height_cm"),
                ),
            )
            shipment_id = int(scur.lastrowid)
            shipment_ids.append(shipment_id)
            for iid in box["item_ids"]:
                item_restores.append(
                    {"id": iid, "status": by_id[iid]["status"]}
                )
                conn.execute(
                    """
                    INSERT INTO shipment_items (shipment_id, item_id, qty)
                    VALUES (?, ?, ?)
                    """,
                    (shipment_id, iid, by_id[iid]["qty"]),
                )
                conn.execute(
                    "UPDATE items SET status = 'outbound_shipped' WHERE id = ?",
                    (iid,),
                )

        for oid in order_ids:
            sync_order_status(conn, oid)

        # Warehouse stock boxes are independent of outbound packing;
        # once orders leave stock, drop them from any stock box.
        release_stock_box_orders(conn, order_ids)

        action_log.record(
            conn,
            "create_outbound_batch",
            f"出库批次 #{batch_id}（{len(normalized)} 箱 / {len(seen_items)} 行）",
            {
                "batch_id": batch_id,
                "shipment_ids": shipment_ids,
                "item_restores": item_restores,
                "order_restores": order_restores,
            },
        )
        return _batch_out(conn, batch_id)


def update_finance(batch_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    """Update international freight fields and/or amount_received_cny."""
    data = {k: v for k, v in payload.items() if v is not None or k in ("payment_note",)}
    if not data:
        return get_batch(batch_id)

    with get_conn() as conn:
        batch = conn.execute(
            "SELECT * FROM outbound_batches WHERE id = ?", (batch_id,)
        ).fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="outbound batch not found")

        freight_rate = (
            float(data["freight_exchange_rate"])
            if "freight_exchange_rate" in data
            else _as_float(batch["freight_exchange_rate"])
        )
        freight_unit = (
            float(data["freight_unit_price_jpy"])
            if "freight_unit_price_jpy" in data
            else _as_float(batch["freight_unit_price_jpy"])
        )
        freight_weight = (
            float(data["chargeable_weight"])
            if "chargeable_weight" in data
            else _as_float(batch["chargeable_weight"])
        )
        if freight_rate is not None and freight_rate <= 0:
            raise HTTPException(
                status_code=400, detail="freight_exchange_rate must be > 0"
            )
        for label, value in (
            ("freight_unit_price_jpy", freight_unit),
            ("chargeable_weight", freight_weight),
        ):
            if value is not None and value < 0:
                raise HTTPException(status_code=400, detail=f"{label} invalid")

        freight_cny = _compute_freight_cny(freight_unit, freight_weight, freight_rate)
        goods_cny = _as_float(batch["goods_receivable_cny"])
        received = (
            float(data["amount_received_cny"])
            if "amount_received_cny" in data
            else (_as_float(batch["amount_received_cny"]) or 0.0)
        )
        if received < 0:
            raise HTTPException(status_code=400, detail="amount_received_cny invalid")
        receivable, pay_status, _ = _recompute_batch_totals(
            goods_cny, freight_cny, received
        )
        payment_note = (
            str(data["payment_note"]).strip()
            if "payment_note" in data
            else (batch["payment_note"] or "")
        )
        inv_date = (
            str(data["invoice_ship_date"]).strip() or None
            if "invoice_ship_date" in data
            else (
                (batch["invoice_ship_date"] if "invoice_ship_date" in batch.keys() else None)
                or None
            )
        )

        conn.execute(
            """
            UPDATE outbound_batches SET
                freight_exchange_rate = ?,
                freight_unit_price_jpy = ?,
                chargeable_weight = ?,
                freight_cny = ?,
                amount_receivable_cny = ?,
                amount_received_cny = ?,
                payment_status = ?,
                payment_note = ?,
                invoice_ship_date = ?
            WHERE id = ?
            """,
            (
                freight_rate,
                freight_unit,
                freight_weight,
                freight_cny,
                receivable,
                received,
                pay_status,
                payment_note,
                inv_date,
                batch_id,
            ),
        )
        action_log.record(
            conn,
            "update_outbound_finance",
            f"更新出库批次 #{batch_id} 财务（{pay_status}）",
            {
                "batch_id": batch_id,
                "before": {
                    "freight_exchange_rate": batch["freight_exchange_rate"],
                    "freight_unit_price_jpy": batch["freight_unit_price_jpy"],
                    "chargeable_weight": batch["chargeable_weight"],
                    "freight_cny": batch["freight_cny"],
                    "amount_receivable_cny": batch["amount_receivable_cny"],
                    "amount_received_cny": batch["amount_received_cny"],
                    "payment_status": batch["payment_status"],
                    "payment_note": batch["payment_note"],
                },
            },
        )
        return _batch_out(conn, batch_id)


def update_batch(
    batch_id: int,
    boxes: list[dict[str, Any]],
    note: Optional[str] = None,
    *,
    allow_missing_barcode: bool = False,
    missing_barcode_note: str = "",
    invoice_ship_date: Optional[str] = None,
) -> dict[str, Any]:
    """
    Re-edit an outbound batch that is not yet fully signed (no delivered boxes).
    Can change box meta, item membership, and item quantities; recalc goods lock.
    """
    if not boxes:
        raise HTTPException(status_code=400, detail="boxes required")
    special_note = (missing_barcode_note or "").strip()
    if allow_missing_barcode and not special_note:
        raise HTTPException(
            status_code=400,
            detail="勾选特殊情况时必须填写无条形码备注",
        )

    normalized: list[dict[str, Any]] = []
    seen_items: set[int] = set()
    qty_by_item: dict[int, int] = {}
    used_box_nos: set[int] = set()
    batch_tracking = _require_shared_batch_tracking(boxes)

    for i, box in enumerate(boxes, start=1):
        raw_items = list(box.get("items") or [])
        if not raw_items and box.get("item_ids"):
            raw_items = [{"item_id": iid, "qty": None} for iid in box["item_ids"]]
        if not raw_items:
            raise HTTPException(status_code=400, detail=f"第{i}箱没有明细行")

        item_ids: list[int] = []
        for entry in raw_items:
            iid = int(entry.get("item_id") or entry.get("id") or 0)
            if iid < 1:
                raise HTTPException(status_code=400, detail=f"第{i}箱明细无效")
            if iid in seen_items:
                raise HTTPException(
                    status_code=400, detail=f"明细行 #{iid} 被分到多个箱"
                )
            seen_items.add(iid)
            item_ids.append(iid)
            qty_raw = entry.get("qty")
            if qty_raw is not None:
                qty = int(qty_raw)
                if qty < 1:
                    raise HTTPException(
                        status_code=400, detail=f"明细行 #{iid} 数量必须 ≥ 1"
                    )
                qty_by_item[iid] = qty

        carrier = (box.get("carrier") or "other").strip().lower()
        if carrier not in ("yamato", "sagawa", "other"):
            raise HTTPException(status_code=400, detail=f"第{i}箱承运商无效")

        box_no = int(box.get("box_no") or i)
        if box_no < 1:
            raise HTTPException(status_code=400, detail=f"第{i}箱箱号无效")
        if box_no in used_box_nos:
            raise HTTPException(status_code=400, detail=f"箱号重复: {box_no}")
        used_box_nos.add(box_no)

        normalized.append(
            {
                "box_no": box_no,
                "carrier": carrier,
                "tracking_no": batch_tracking,
                "note": (box.get("note") or "").strip(),
                "item_ids": item_ids,
                "net_weight": _as_float(box.get("net_weight")),
                "gross_weight": _as_float(box.get("gross_weight")),
                "length_cm": _as_float(box.get("length_cm")),
                "width_cm": _as_float(box.get("width_cm")),
                "height_cm": _as_float(box.get("height_cm")),
            }
        )

    with get_conn() as conn:
        batch = conn.execute(
            "SELECT * FROM outbound_batches WHERE id = ?", (batch_id,)
        ).fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="outbound batch not found")

        ships = conn.execute(
            """
            SELECT id, status, tracking_no FROM shipments
            WHERE batch_id = ? AND direction = 'outbound'
            """,
            (batch_id,),
        ).fetchall()
        if not ships:
            raise HTTPException(status_code=400, detail="批次没有出库箱")
        if any((s["status"] or "") == "delivered" for s in ships):
            raise HTTPException(
                status_code=400,
                detail="已有箱子签收，不能再编辑该批次",
            )

        old_rows = conn.execute(
            """
            SELECT si.item_id, si.qty, i.order_id, i.status
            FROM shipment_items si
            JOIN shipments s ON s.id = si.shipment_id
            JOIN items i ON i.id = si.item_id
            WHERE s.batch_id = ? AND s.direction = 'outbound'
            """,
            (batch_id,),
        ).fetchall()
        old_item_ids = {int(r["item_id"]) for r in old_rows}
        old_order_ids = {int(r["order_id"]) for r in old_rows if r["order_id"]}

        exists = conn.execute(
            """
            SELECT id FROM shipments
            WHERE tracking_no = ?
              AND NOT (batch_id = ? AND direction = 'outbound')
            """,
            (batch_tracking, batch_id),
        ).fetchone()
        if exists:
            raise HTTPException(
                status_code=409,
                detail=f"运单号已被其他出库批次或进库使用: {batch_tracking}",
            )

        placeholders = ",".join("?" * len(seen_items))
        rows = conn.execute(
            f"""
            SELECT id, order_id, order_ref, status, qty, name, barcode
            FROM items WHERE id IN ({placeholders})
            """,
            list(seen_items),
        ).fetchall()
        by_id = {int(r["id"]): r for r in rows}
        if len(by_id) != len(seen_items):
            missing = [iid for iid in seen_items if iid not in by_id]
            raise HTTPException(status_code=404, detail=f"items not found: {missing}")

        for iid, row in by_id.items():
            st = row["status"]
            if iid in old_item_ids:
                if st not in ("outbound_shipped", "in_stock"):
                    raise HTTPException(
                        status_code=400,
                        detail=f"明细 #{iid} 状态不可编辑: {st}",
                    )
            else:
                if st != "in_stock":
                    raise HTTPException(
                        status_code=400,
                        detail=f"明细 #{iid} 不在库，不能加入批次",
                    )

        no_barcode = [
            r for r in by_id.values() if not (r["barcode"] or "").strip()
        ]
        if no_barcode and not allow_missing_barcode:
            names = "、".join(f"「{r['name']}」" for r in no_barcode[:5])
            more = f" 等{len(no_barcode)}件" if len(no_barcode) > 5 else ""
            raise HTTPException(
                status_code=400,
                detail=f"出库前须登记条形码：{names}{more}",
            )

        order_ids = {int(r["order_id"]) for r in by_id.values() if r["order_id"]}
        # Full-order rule among in_stock + this batch
        for oid in order_ids:
            eligible = conn.execute(
                """
                SELECT id FROM items
                WHERE order_id = ?
                  AND (
                    status = 'in_stock'
                    OR id IN (
                      SELECT si.item_id FROM shipment_items si
                      JOIN shipments s ON s.id = si.shipment_id
                      WHERE s.batch_id = ? AND s.direction = 'outbound'
                    )
                  )
                """,
                (oid, batch_id),
            ).fetchall()
            eligible_ids = {int(r["id"]) for r in eligible}
            missing = eligible_ids - seen_items
            if missing:
                ref = conn.execute(
                    "SELECT order_ref FROM orders WHERE id = ?", (oid,)
                ).fetchone()
                label = (ref["order_ref"] if ref else "") or f"#{oid}"
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"订单 {label} 不能部分出库，未装箱的明细: "
                        f"{sorted(missing)}"
                    ),
                )

        # Apply qty updates before rebuild
        for iid, qty in qty_by_item.items():
            conn.execute("UPDATE items SET qty = ? WHERE id = ?", (qty, iid))

        # Release removed items back to stock
        removed = old_item_ids - seen_items
        for iid in removed:
            conn.execute(
                "UPDATE items SET status = 'in_stock' WHERE id = ?", (iid,)
            )

        # Delete old outbound shipments (shipment_items cascade)
        for ship in ships:
            conn.execute("DELETE FROM shipment_items WHERE shipment_id = ?", (ship["id"],))
            conn.execute("DELETE FROM shipments WHERE id = ?", (ship["id"],))

        # Refresh qty after updates
        rows2 = conn.execute(
            f"""
            SELECT id, order_id, order_ref, status, qty, name, barcode
            FROM items WHERE id IN ({placeholders})
            """,
            list(seen_items),
        ).fetchall()
        by_id = {int(r["id"]): r for r in rows2}

        for box in sorted(normalized, key=lambda b: b["box_no"]):
            scur = conn.execute(
                """
                INSERT INTO shipments (
                    direction, carrier, tracking_no, shipped_at, status,
                    batch_id, box_no, note,
                    net_weight, gross_weight, length_cm, width_cm, height_cm
                ) VALUES ('outbound', ?, ?, ?, 'shipped', ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    box["carrier"],
                    box["tracking_no"],
                    _now(),
                    batch_id,
                    box["box_no"],
                    box.get("note") or "",
                    box.get("net_weight"),
                    box.get("gross_weight"),
                    box.get("length_cm"),
                    box.get("width_cm"),
                    box.get("height_cm"),
                ),
            )
            shipment_id = int(scur.lastrowid)
            for iid in box["item_ids"]:
                conn.execute(
                    """
                    INSERT INTO shipment_items (shipment_id, item_id, qty)
                    VALUES (?, ?, ?)
                    """,
                    (shipment_id, iid, by_id[iid]["qty"]),
                )
                conn.execute(
                    "UPDATE items SET status = 'outbound_shipped' WHERE id = ?",
                    (iid,),
                )

        locked = _lock_receivables(conn, seen_items, order_ids)
        freight_rate = _as_float(batch["freight_exchange_rate"])
        freight_unit = _as_float(batch["freight_unit_price_jpy"])
        freight_weight = _as_float(batch["chargeable_weight"])
        freight_cny = _compute_freight_cny(freight_unit, freight_weight, freight_rate)
        received = _as_float(batch["amount_received_cny"]) or 0.0
        receivable, pay_status, _ = _recompute_batch_totals(
            locked["goods_receivable_cny"], freight_cny, received
        )

        batch_note = batch["note"] or ""
        if note is not None:
            batch_note = note.strip()
        if allow_missing_barcode and no_barcode:
            tag = f"【无条码特批】{special_note}"
            if "【无条码特批】" not in batch_note:
                batch_note = f"{batch_note} | {tag}".strip(" |") if batch_note else tag

        ship_date = (
            (invoice_ship_date or "").strip()
            if invoice_ship_date is not None
            else (
                (batch["invoice_ship_date"] if "invoice_ship_date" in batch.keys() else None)
                or None
            )
        )

        conn.execute(
            """
            UPDATE outbound_batches SET
                note = ?,
                goods_jpy = ?,
                order_shipping_jpy = ?,
                goods_receivable_cny = ?,
                freight_cny = ?,
                amount_receivable_cny = ?,
                payment_status = ?,
                invoice_ship_date = ?
            WHERE id = ?
            """,
            (
                batch_note,
                locked["goods_jpy"],
                locked["order_shipping_jpy"],
                locked["goods_receivable_cny"],
                freight_cny,
                receivable,
                pay_status,
                ship_date or None,
                batch_id,
            ),
        )

        touched_orders = old_order_ids | order_ids
        for oid in touched_orders:
            sync_order_status(conn, oid)
        release_stock_box_orders(conn, order_ids)

        action_log.record(
            conn,
            "update_outbound_batch",
            f"编辑出库批次 #{batch_id}（{len(normalized)} 箱 / {len(seen_items)} 行）",
            {"batch_id": batch_id, "item_ids": sorted(seen_items)},
        )
        return _batch_out(conn, batch_id)


def confirm_batch(batch_id: int) -> dict[str, Any]:
    """Confirm all shipped boxes in a batch as delivered (user signed)."""
    with get_conn() as conn:
        batch = conn.execute(
            "SELECT id FROM outbound_batches WHERE id = ?", (batch_id,)
        ).fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="outbound batch not found")
        ships = conn.execute(
            """
            SELECT id FROM shipments
            WHERE batch_id = ? AND direction = 'outbound' AND status = 'shipped'
            ORDER BY box_no
            """,
            (batch_id,),
        ).fetchall()
        for ship in ships:
            # inline confirm to share connection
            shipment_id = ship["id"]
            now = _now()
            item_rows = conn.execute(
                """
                SELECT i.id, i.order_id, i.status, i.arrived_at
                FROM shipment_items si
                JOIN items i ON i.id = si.item_id
                WHERE si.shipment_id = ?
                """,
                (shipment_id,),
            ).fetchall()
            order_ids: set[int] = set()
            for r in item_rows:
                if r["status"] == "cancelled":
                    continue
                conn.execute(
                    "UPDATE items SET status = 'delivered', arrived_at = ? WHERE id = ?",
                    (now, r["id"]),
                )
                if r["order_id"] is not None:
                    order_ids.add(int(r["order_id"]))
            conn.execute(
                """
                UPDATE shipments
                SET status = 'delivered', delivered_at = ?
                WHERE id = ?
                """,
                (now, shipment_id),
            )
            for oid in order_ids:
                sync_order_status(conn, oid)

        if ships:
            action_log.record(
                conn,
                "confirm_outbound_batch",
                f"确认出库批次 #{batch_id} 签收（{len(ships)} 箱）",
                {"batch_id": batch_id, "shipment_ids": [s["id"] for s in ships]},
            )
        return _batch_out(conn, batch_id)


def export_fee_detail_xlsx(batch_id: int) -> bytes:
    """Export 发货费用明细 Excel (template: 发货费用明细 + 对应订单)."""
    template = DATA_DIR / "templates" / "fee_detail.xlsx"
    if not template.is_file():
        raise HTTPException(
            status_code=500, detail=f"fee-detail template missing: {template}"
        )

    with get_conn() as conn:
        batch = _batch_out(conn, batch_id)
        lines: list[dict[str, Any]] = []
        order_qty: dict[tuple[int, str], dict[str, Any]] = {}
        for box in batch["boxes"]:
            tracking = (box.get("tracking_no") or "").strip()
            box_no = int(box["box_no"])
            for item in box["items"]:
                item_row = conn.execute(
                    """
                    SELECT i.id, i.name, i.barcode, i.qty, i.unit_cost,
                           i.order_ref, o.order_ref AS order_order_ref,
                           o.exchange_rate
                    FROM items i
                    JOIN orders o ON o.id = i.order_id
                    WHERE i.id = ?
                    """,
                    (item["id"],),
                ).fetchone()
                if not item_row:
                    continue
                unit = _as_float(item_row["unit_cost"])
                qty = int(item_row["qty"])
                amount_jpy = round(unit * qty, 2) if unit is not None else None
                rate = _as_float(item_row["exchange_rate"])
                if rate is not None and rate <= 0:
                    rate = None
                amount_cny = (
                    round(amount_jpy * rate, 2)
                    if amount_jpy is not None and rate is not None
                    else None
                )
                order_ref = (
                    (item_row["order_order_ref"] or item_row["order_ref"] or "").strip()
                    or f"#{item.get('order_id') or ''}"
                )
                lines.append(
                    {
                        "box_no": box_no,
                        "order_ref": order_ref,
                        "name": item_row["name"],
                        "barcode": item_row["barcode"] or "",
                        "qty": qty,
                        "amount_jpy": amount_jpy,
                        "exchange_rate": rate,
                        "amount_cny": amount_cny,
                    }
                )
                key = (box_no, order_ref)
                if key not in order_qty:
                    order_qty[key] = {
                        "box_no": box_no,
                        "tracking_no": tracking,
                        "order_ref": order_ref,
                        "qty": 0,
                    }
                order_qty[key]["qty"] += qty

    order_rows = sorted(
        order_qty.values(),
        key=lambda x: (int(x["box_no"]), str(x["order_ref"])),
    )

    wb = load_workbook(template)
    if "发货费用明细" not in wb.sheetnames or "对应订单" not in wb.sheetnames:
        raise HTTPException(
            status_code=500,
            detail=f"fee-detail template sheets missing: {wb.sheetnames}",
        )
    ws = wb["发货费用明细"]
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True)

    invoice_no = f"OB-{batch_id}"
    for i, line in enumerate(lines, start=1):
        r = i + 1
        values = [
            invoice_no if i == 1 else None,
            line["box_no"],
            line["order_ref"],
            line["name"],
            line["barcode"],
            line["qty"],
            line["amount_jpy"],
            line["exchange_rate"],
            line["amount_cny"],
            None,
            None,
            None,
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(r, c, v)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    last_data_row = 1 + len(lines)
    if len(lines) > 1:
        ws.merge_cells(start_row=2, start_column=1, end_row=last_data_row, end_column=1)
        ws.cell(2, 1).alignment = Alignment(vertical="center", wrap_text=True)

    total_row = last_data_row + 1
    total_qty = sum(int(x["qty"]) for x in lines)
    total_jpy = sum(float(x["amount_jpy"] or 0) for x in lines)
    total_cny = sum(float(x["amount_cny"] or 0) for x in lines)
    ws.cell(total_row, 6, total_qty).border = thin
    ws.cell(total_row, 7, round(total_jpy, 2)).border = thin
    ws.cell(total_row, 9, round(total_cny, 2) if any(x["amount_cny"] is not None for x in lines) else None).border = thin
    ws.cell(total_row, 6).font = header_font
    ws.cell(total_row, 7).font = header_font
    ws.cell(total_row, 9).font = header_font

    rates = {x["exchange_rate"] for x in lines if x["exchange_rate"] is not None}
    goods_rate_summary: Any
    if len(rates) == 1:
        goods_rate_summary = next(iter(rates))
    elif len(rates) > 1:
        goods_rate_summary = "分订单见各行"
    else:
        goods_rate_summary = None

    summary_rows = [
        ("商品汇率", goods_rate_summary),
        ("运费汇率", batch.get("freight_exchange_rate")),
        ("计费重量", batch.get("chargeable_weight")),
        ("运费单价（JPY）", batch.get("freight_unit_price_jpy")),
        ("需支付运费（CNY）", batch.get("freight_cny")),
        ("需支付货款（CNY）", batch.get("goods_receivable_cny")),
        ("总支付金额（CNY）", batch.get("amount_receivable_cny")),
    ]
    for i, (label, value) in enumerate(summary_rows):
        r = total_row + 1 + i
        ws.cell(r, 6, label)
        ws.cell(r, 7, value)

    for c, w in enumerate([14, 8, 16, 28, 16, 8, 12, 10, 12, 8, 8, 12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    orders_ws = wb["对应订单"]
    row_i = 2
    groups: list[tuple[int, int, int]] = []  # box_no, start, end
    prev_box: Optional[int] = None
    group_start = 2
    for rec in order_rows:
        box_no = int(rec["box_no"])
        if prev_box is not None and box_no != prev_box:
            groups.append((prev_box, group_start, row_i - 1))
            group_start = row_i
        orders_ws.cell(row_i, 1, box_no).border = thin
        orders_ws.cell(row_i, 2, rec.get("tracking_no") or "").border = thin
        orders_ws.cell(row_i, 3, rec["order_ref"]).border = thin
        orders_ws.cell(row_i, 4, int(rec["qty"] or 0)).border = thin
        prev_box = box_no
        row_i += 1
    if prev_box is not None:
        groups.append((prev_box, group_start, row_i - 1))
    for _box_no, start, end in groups:
        if end > start:
            orders_ws.merge_cells(
                start_row=start, start_column=1, end_row=end, end_column=1
            )
            orders_ws.merge_cells(
                start_row=start, start_column=2, end_row=end, end_column=2
            )
        orders_ws.cell(start, 1).alignment = Alignment(
            vertical="center", horizontal="center"
        )
        orders_ws.cell(start, 2).alignment = Alignment(vertical="center")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _require_box_packing(box_no: int, box: dict[str, Any]) -> None:
    missing: list[str] = []
    for key, label in (
        ("net_weight", "净重"),
        ("gross_weight", "毛重"),
        ("length_cm", "长"),
        ("width_cm", "宽"),
        ("height_cm", "高"),
    ):
        if _as_float(box.get(key)) is None:
            missing.append(label)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"第{box_no}箱请填写完整包装信息：{'/'.join(missing)}",
        )


def export_inv_xlsx(batch_id: int) -> tuple[bytes, str]:
    """Export INV from fixed FIT dual-sheet template. Returns (bytes, filename)."""
    from app.services.inv_template import (
        build_inv_workbook,
        dim_text,
        kind_to_en,
        parse_ship_date,
    )

    with get_conn() as conn:
        batch = _batch_out(conn, batch_id)
        product_lines: list[dict[str, Any]] = []
        packing_lines: list[dict[str, Any]] = []
        for box in batch["boxes"]:
            kind_counts: dict[str, int] = {}
            box_qty = 0
            for item in box["items"]:
                item_row = conn.execute(
                    """
                    SELECT i.barcode, i.qty, i.unit_cost, i.product_kind
                    FROM items i WHERE i.id = ?
                    """,
                    (item["id"],),
                ).fetchone()
                if not item_row:
                    continue
                qty = int(item_row["qty"] or item.get("qty") or 0)
                kind = kind_to_en(item_row["product_kind"] or "")
                product_lines.append(
                    {
                        "barcode": item_row["barcode"] or "",
                        "classify_en": kind,
                        "qty": qty,
                        "unit_price": _as_float(item_row["unit_cost"]),
                    }
                )
                kind_counts[kind] = kind_counts.get(kind, 0) + qty
                box_qty += qty
            if box_qty <= 0:
                continue
            _require_box_packing(int(box["box_no"]), box)
            ordered = sorted(kind_counts.items(), key=lambda kv: (-kv[1], kv[0]))
            commodity = " / ".join(k for k, _ in ordered) if ordered else "Toys"
            packing_lines.append(
                {
                    "packing_no": f"FIT{batch_id}-{int(box['box_no']):02d}",
                    "commodity_en": commodity,
                    "qty": box_qty,
                    "net_weight": box.get("net_weight"),
                    "gross_weight": box.get("gross_weight"),
                    "dim_text": dim_text(
                        box.get("length_cm"),
                        box.get("width_cm"),
                        box.get("height_cm"),
                    ),
                }
            )
        ship_date = parse_ship_date(batch.get("invoice_ship_date"))
        content, _inv_no, filename = build_inv_workbook(
            batch_id=batch_id,
            ship_date=ship_date,
            product_lines=product_lines,
            packing_lines=packing_lines,
        )
        return content, filename


def export_inv_preview_xlsx(payload: dict[str, Any]) -> tuple[bytes, str]:
    """Draft INV preview using batch_id=0 in INV number."""
    from app.services.inv_template import (
        build_inv_workbook,
        dim_text,
        kind_to_en,
        parse_ship_date,
    )

    boxes = list(payload.get("boxes") or [])
    if not boxes:
        raise HTTPException(status_code=400, detail="boxes required")
    item_ids: list[int] = []
    for i, box in enumerate(boxes, start=1):
        ids = [int(x) for x in (box.get("item_ids") or [])]
        if not ids:
            raise HTTPException(status_code=400, detail=f"第{i}箱没有明细行")
        item_ids.extend(ids)
    if len(set(item_ids)) != len(item_ids):
        raise HTTPException(status_code=400, detail="明细行不能重复分箱")

    with get_conn() as conn:
        placeholders = ",".join("?" * len(item_ids))
        rows = conn.execute(
            f"""
            SELECT id, barcode, qty, unit_cost, product_kind
            FROM items WHERE id IN ({placeholders})
            """,
            item_ids,
        ).fetchall()
        by_id = {int(r["id"]): r for r in rows}
        if len(by_id) != len(set(item_ids)):
            missing = [iid for iid in set(item_ids) if iid not in by_id]
            raise HTTPException(status_code=404, detail=f"items not found: {missing}")

        product_lines: list[dict[str, Any]] = []
        packing_lines: list[dict[str, Any]] = []
        for i, box in enumerate(boxes, start=1):
            box_no = int(box.get("box_no") or i)
            kind_counts: dict[str, int] = {}
            box_qty = 0
            for iid in box.get("item_ids") or []:
                row = by_id[int(iid)]
                qty = int(row["qty"] or 0)
                kind = kind_to_en(row["product_kind"] or "")
                product_lines.append(
                    {
                        "barcode": row["barcode"] or "",
                        "classify_en": kind,
                        "qty": qty,
                        "unit_price": _as_float(row["unit_cost"]),
                    }
                )
                kind_counts[kind] = kind_counts.get(kind, 0) + qty
                box_qty += qty
            _require_box_packing(
                box_no,
                {
                    "net_weight": box.get("net_weight"),
                    "gross_weight": box.get("gross_weight"),
                    "length_cm": box.get("length_cm"),
                    "width_cm": box.get("width_cm"),
                    "height_cm": box.get("height_cm"),
                },
            )
            ordered = sorted(kind_counts.items(), key=lambda kv: (-kv[1], kv[0]))
            commodity = " / ".join(k for k, _ in ordered) if ordered else "Toys"
            packing_lines.append(
                {
                    "packing_no": f"DRAFT-{box_no:02d}",
                    "commodity_en": commodity,
                    "qty": box_qty,
                    "net_weight": _as_float(box.get("net_weight")),
                    "gross_weight": _as_float(box.get("gross_weight")),
                    "dim_text": dim_text(
                        _as_float(box.get("length_cm")),
                        _as_float(box.get("width_cm")),
                        _as_float(box.get("height_cm")),
                    ),
                }
            )
        ship_date = parse_ship_date(payload.get("invoice_ship_date"))
        content, _inv_no, filename = build_inv_workbook(
            batch_id=0,
            ship_date=ship_date,
            product_lines=product_lines,
            packing_lines=packing_lines,
        )
        return content, filename
