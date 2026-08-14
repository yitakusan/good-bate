"""Fill FIT INV shipping Excel from the fixed dual-sheet template."""

from __future__ import annotations

from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.database import DATA_DIR

# ============================================================
# FEATURE: INV_EXPORT
#
# [用途] 填充固定双 Sheet 模板（inv + PACKING LIST ）
# [模板] backend/data/templates/inv_fit_shipping.xlsx（只读）
# [调用] outbound_batches.export_inv_xlsx / export_inv_preview_xlsx
# [代码索引] docs/CODE_INDEX.md#feature-inv_export
# ============================================================

TEMPLATE_PATH = DATA_DIR / "templates" / "inv_fit_shipping.xlsx"
INV_SHEET = "inv"
PACK_SHEET = "PACKING LIST "

CONSIGNEE_EN = (
    "Zhang Chao\n"
    "Tengyue Cloud Warehouse, No. 8 Hongfa Road, Development Zone, "
    "Cuihuangkou Town, Wuqing District, Tianjin, China\n"
    "TEL: 15033057065\n"
    "POST CODE: 301702"
)

KIND_EN: dict[str, str] = {
    "手办": "Figures",
    "模型": "Models",
    "毛绒": "Plush Toys",
    "毛绒玩具": "Plush Toys",
    "玩偶": "Plush Toys",
    "徽章": "Badges",
    "吧唧": "Badges",
    "亚克力立牌": "Acrylic Stands",
    "立牌": "Acrylic Stands",
    "钥匙扣": "Keychains",
    "挂件": "Keychains",
    "卡牌": "Trading Cards",
    "小卡": "Trading Cards",
    "玩具": "Toys",
    "海报": "Posters",
}


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def parse_ship_date(value: Optional[str]) -> date:
    raw = (value or "").strip()
    if not raw:
        return date.today()
    cleaned = raw[:10].replace("/", "-")
    try:
        return datetime.strptime(cleaned, "%Y-%m-%d").date()
    except ValueError:
        return date.today()


def format_inv_no(ship_date: date, batch_id: int, total_qty: int) -> str:
    return f"{ship_date.year:04d}{ship_date.month:02d}/{ship_date.day:02d}FIT{batch_id}-{total_qty}"


def format_filename(ship_date: date, batch_id: int, total_qty: int) -> str:
    return f"INV_{ship_date.year:04d}{ship_date.month:02d}{ship_date.day:02d}_FIT{batch_id}_{total_qty}.xlsx"


def kind_to_en(kind: str) -> str:
    text = (kind or "").strip()
    if not text:
        return "Toys"
    if text in KIND_EN:
        return KIND_EN[text]
    for cn, en in KIND_EN.items():
        if cn in text:
            return en
    # already English-ish
    if text.isascii():
        return text
    return "Toys"


def dim_text(length: Optional[float], width: Optional[float], height: Optional[float]) -> str:
    if length is None or width is None or height is None:
        return ""
    def fmt(v: float) -> str:
        return str(int(v)) if float(v).is_integer() else str(v)
    return f"{fmt(length)}*{fmt(width)}*{fmt(height)}"


def _ensure_template() -> Path:
    if not TEMPLATE_PATH.is_file():
        raise HTTPException(
            status_code=500,
            detail=f"INV template missing: {TEMPLATE_PATH}",
        )
    return TEMPLATE_PATH


# ============================================================
# FEATURE: INV_EXPORT
# [业务逻辑] build_inv_workbook — 填充模板双 Sheet
# ============================================================
def build_inv_workbook(
    *,
    batch_id: int,
    ship_date: date,
    product_lines: list[dict[str, Any]],
    packing_lines: list[dict[str, Any]],
) -> tuple[bytes, str, str]:
    """
    product_lines: barcode, classify_en, qty, unit_price
    packing_lines: packing_no, commodity_en, qty, net_weight, gross_weight, dim_text
    """
    total_qty = sum(int(p.get("qty") or 0) for p in product_lines)
    if total_qty < 1:
        raise HTTPException(status_code=400, detail="INV 需要至少 1 件商品")
    inv_no = format_inv_no(ship_date, batch_id, total_qty)
    filename = format_filename(ship_date, batch_id, total_qty)

    wb = load_workbook(_ensure_template())
    if INV_SHEET not in wb.sheetnames or PACK_SHEET not in wb.sheetnames:
        raise HTTPException(
            status_code=500,
            detail=f"INV template sheets missing: {wb.sheetnames}",
        )
    inv = wb[INV_SHEET]
    pack = wb[PACK_SHEET]

    # --- inv sheet ---
    inv["F2"] = ship_date.isoformat()
    inv["F3"] = inv_no
    inv["F4"] = inv_no
    inv["A3"] = CONSIGNEE_EN

    data_start = 8
    template_end = 12
    n = len(product_lines)
    template_count = template_end - data_start + 1
    if n > template_count:
        extra = n - template_count
        inv.insert_rows(template_end + 1, amount=extra)
        for i in range(extra):
            _copy_row_style(inv, data_start, template_end + 1 + i, 7)

    for i, line in enumerate(product_lines):
        r = data_start + i
        inv.cell(r, 1, (line.get("barcode") or "").strip())
        inv.cell(r, 2, line.get("classify_en") or "Toys")
        inv.cell(r, 3, "China")
        inv.cell(r, 4, int(line.get("qty") or 0))
        inv.cell(r, 5, "piece")
        price = line.get("unit_price")
        inv.cell(r, 6, price if price is not None else None)
        inv.cell(r, 7, f"=D{r}*F{r}")

    # clear leftover sample rows
    clear_to = max(template_end, data_start + n - 1)
    for r in range(data_start + n, clear_to + 1):
        for c in range(1, 8):
            inv.cell(r, c, None)

    # --- packing list ---
    pack["F4"] = datetime(ship_date.year, ship_date.month, ship_date.day)
    pack["F5"] = inv_no
    pack["A8"] = CONSIGNEE_EN

    pack_start = 22
    pack_template_end = 25
    total_row = 26
    m = len(packing_lines)
    if m < 1:
        raise HTTPException(status_code=400, detail="INV 需要至少 1 箱包装信息")

    pack_template_count = pack_template_end - pack_start + 1
    if m > pack_template_count:
        extra = m - pack_template_count
        # insert above TOTAL row
        pack.insert_rows(total_row, amount=extra)
        for i in range(extra):
            _copy_row_style(pack, pack_start, total_row + i, 6)
        total_row = pack_start + m

    for i, line in enumerate(packing_lines):
        r = pack_start + i
        pack.cell(r, 1, line.get("packing_no") or "")
        pack.cell(r, 2, line.get("commodity_en") or "Toys")
        pack.cell(r, 3, int(line.get("qty") or 0))
        pack.cell(r, 4, line.get("net_weight"))
        pack.cell(r, 5, line.get("gross_weight"))
        pack.cell(r, 6, line.get("dim_text") or "")

    # clear leftover packing sample rows (before TOTAL)
    for r in range(pack_start + m, pack_start + max(m, pack_template_count)):
        if r >= total_row:
            break
        for c in range(1, 7):
            pack.cell(r, c, None)

    # ensure TOTAL row
    end_data = pack_start + m - 1
    # if we didn't insert, TOTAL stays at 26; if fewer rows, keep TOTAL at 26 and clear middle
    if m <= pack_template_count:
        total_row = 26
        for r in range(pack_start + m, total_row):
            for c in range(1, 7):
                pack.cell(r, c, None)
    pack.cell(total_row, 1, "TOTAL")
    pack.cell(total_row, 3, f"=SUM(C{pack_start}:C{end_data})")
    pack.cell(total_row, 4, f"=SUM(D{pack_start}:D{end_data})")
    pack.cell(total_row, 5, f"=SUM(E{pack_start}:E{end_data})")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), inv_no, filename
